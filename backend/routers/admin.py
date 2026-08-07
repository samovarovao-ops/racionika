import os
from fastapi import APIRouter, HTTPException, Depends, Header
from typing import Optional
from backend.storage import get_menu_data, list_all_menus
from backend.menu_parser import MenuData

router = APIRouter()

ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "admin123")


@router.post("/admin/login")
def admin_login(body: dict):
    password = body.get("password", "")
    if password == ADMIN_PASSWORD:
        return {"status": "ok"}
    raise HTTPException(status_code=401, detail="Неверный пароль")


def verify_admin(x_admin_password: Optional[str] = Header(None)):
    if x_admin_password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="Unauthorized")
    return True


@router.get("/admin/versions")
def admin_versions(authorized: bool = Depends(verify_admin)):
    return list_all_menus()


@router.get("/admin/menus")
def admin_list_menus(authorized: bool = Depends(verify_admin)):
    return list_all_menus()


@router.get("/admin/menu/{menu_id}")
def admin_menu_detail(menu_id: str, authorized: bool = Depends(verify_admin)):
    menu_data = get_menu_data(menu_id)
    if menu_data is None:
        raise HTTPException(status_code=404, detail="Меню не найдено")
    return {
        "menu_id": menu_id,
        "items": [item.model_dump() for item in menu_data.menu_items],
        "tariffs": [t.model_dump() for t in menu_data.tariffs],
        "settings": menu_data.settings,
    }


@router.put("/admin/menu/{menu_id}/settings")
def update_settings(menu_id: str, body: dict, authorized: bool = Depends(verify_admin)):
    from backend.storage import get_menu_file_path
    import openpyxl

    file_path = get_menu_file_path(menu_id)
    if not file_path:
        raise HTTPException(status_code=404, detail="Меню не найдено")

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception:
        raise HTTPException(status_code=500, detail="Ошибка открытия файла")

    if "Настройки" not in wb.sheetnames:
        wb.close()
        raise HTTPException(status_code=400, detail="Лист 'Настройки' не найден")

    ws = wb["Настройки"]
    existing = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            existing[str(row[0])] = row[1]

    for key, value in body.items():
        if key in existing:
            for row in ws.iter_rows(min_row=2, values_only=False):
                if str(row[0].value) == key:
                    row[1].value = str(value)
                    break
        else:
            ws.append([key, str(value)])

    try:
        wb.save(file_path)
        wb.close()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка сохранения: {e}")

    return {"status": "ok", "settings": body}


@router.put("/admin/menu/{menu_id}/tariffs")
def update_tariffs(menu_id: str, body: list[dict], authorized: bool = Depends(verify_admin)):
    from backend.storage import get_menu_file_path
    import openpyxl

    file_path = get_menu_file_path(menu_id)
    if not file_path:
        raise HTTPException(status_code=404, detail="Меню не найдено")

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception:
        raise HTTPException(status_code=500, detail="Ошибка открытия файла")

    if "Тарифы" not in wb.sheetnames:
        wb.close()
        raise HTTPException(status_code=400, detail="Лист 'Тарифы' не найден")

    ws = wb["Тарифы"]
    while ws.max_row > 1:
        ws.delete_rows(2)

    for tariff in body:
        ws.append([
            tariff.get("Программа", ""),
            tariff.get("Люди", 0),
            tariff.get("Цена_за_неделю", 0),
        ])

    try:
        wb.save(file_path)
        wb.close()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Ошибка сохранения: {e}")

    return {"status": "ok", "tariffs": body}
