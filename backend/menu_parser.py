import openpyxl
from typing import Optional
from backend.models.schemas import MenuItem, Tariff, Setting

REQUIRED_MENU_COLUMNS = [
    "Программа", "День", "Приём", "Блюдо_ID", "Название_блюда",
    "Описание", "Порции_на_человека", "Цена_за_порцию", "Вес_грамм", "Теги"
]
REQUIRED_TARIFF_COLUMNS = ["Программа", "Люди", "Цена_за_неделю"]
REQUIRED_SETTING_COLUMNS = ["Ключ", "Значение"]
VALID_PROGRAMS = {"Classic", "Balance", "Vegan"}
VALID_MEALS = {"Завтрак", "Обед", "Ужин", "Перекус"}


class ParseError:
    def __init__(self, sheet: str, row: int, column: str, message: str):
        self.sheet = sheet
        self.row = row
        self.column = column
        self.message = message

    def __str__(self):
        return f"[{self.sheet}] строка {self.row}, столбец '{self.column}': {self.message}"

    def to_dict(self):
        return {"sheet": self.sheet, "row": self.row, "column": self.column, "message": self.message}


class MenuData:
    def __init__(self):
        self.menu_items: list[MenuItem] = []
        self.tariffs: list[Tariff] = []
        self.settings: dict[str, str] = {}
        self.errors: list[ParseError] = []
        self.warnings: list[str] = []


def parse_menu(file_path: str) -> MenuData:
    data = MenuData()
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        data.errors.append(ParseError("Файл", 0, "", f"Не удалось открыть файл: {e}"))
        return data

    required_sheets = {"Меню", "Тарифы", "Настройки"}
    found_sheets = set(wb.sheetnames)
    missing = required_sheets - found_sheets
    if missing:
        for s in missing:
            data.errors.append(ParseError("Файл", 0, "", f"Отсутствует лист '{s}'"))

    if "Меню" in wb.sheetnames:
        _parse_menu_sheet(wb["Меню"], data)
    if "Тарифы" in wb.sheetnames:
        _parse_tariff_sheet(wb["Тарифы"], data)
    if "Настройки" in wb.sheetnames:
        _parse_settings_sheet(wb["Настройки"], data)

    wb.close()
    return data


def _validate_headers(ws, required: list[str], sheet_name: str, data: MenuData):
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value else "")
    for col_name in required:
        if col_name not in headers:
            data.errors.append(ParseError(sheet_name, 1, col_name, f"Отсутствует обязательный столбец '{col_name}'"))
    return headers


def _parse_menu_sheet(ws, data: MenuData):
    headers = _validate_headers(ws, REQUIRED_MENU_COLUMNS, "Меню", data)
    if data.errors:
        return

    col_idx = {h: i for i, h in enumerate(headers)}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        vals = list(row)
        if all(v is None for v in vals):
            continue

        def get(col, default=None):
            idx = col_idx.get(col)
            if idx is None or idx >= len(vals):
                return default
            return vals[idx] if vals[idx] is not None else default

        program = str(get("Программа", "")).strip()
        if program not in VALID_PROGRAMS:
            data.errors.append(ParseError("Меню", row_num, "Программа",
                                          f"Недопустимое значение '{program}'. Допустимы: {', '.join(sorted(VALID_PROGRAMS))}"))
            continue

        try:
            day = int(get("День", 0))
        except (ValueError, TypeError):
            data.errors.append(ParseError("Меню", row_num, "День", "Значение должно быть целым числом"))
            continue
        if day < 1 or day > 7:
            data.errors.append(ParseError("Меню", row_num, "День", f"День должен быть 1..7, получено {day}"))
            continue

        meal = str(get("Приём", "")).strip()
        if meal not in VALID_MEALS:
            data.errors.append(ParseError("Меню", row_num, "Приём",
                                          f"Недопустимое значение '{meal}'. Допустимы: {', '.join(sorted(VALID_MEALS))}"))
            continue

        name = str(get("Название_блюда", "")).strip()
        if not name:
            data.errors.append(ParseError("Меню", row_num, "Название_блюда", "Пустое название блюда"))
            continue

        try:
            portions = float(get("Порции_на_человека", 1))
            if portions <= 0:
                raise ValueError()
        except (ValueError, TypeError):
            data.errors.append(ParseError("Меню", row_num, "Порции_на_человека", "Значение должно быть > 0"))
            continue

        try:
            price = float(get("Цена_за_порцию", 0))
            if price < 0:
                raise ValueError()
        except (ValueError, TypeError):
            data.errors.append(ParseError("Меню", row_num, "Цена_за_порцию", "Значение должно быть >= 0"))
            continue

        weight = get("Вес_грамм")
        if weight is not None:
            try:
                weight = float(weight)
            except (ValueError, TypeError):
                weight = None

        data.menu_items.append(MenuItem(
            Программа=program,
            День=day,
            Приём=meal,
            Блюдо_ID=str(get("Блюдо_ID", "")) or None,
            Название_блюда=name,
            Описание=str(get("Описание", "")) or "",
            Порции_на_человека=portions,
            Цена_за_порцию=price,
            Вес_грамм=weight,
            Теги=str(get("Теги", "")) or None,
        ))

    if not data.menu_items:
        data.errors.append(ParseError("Меню", 0, "", "Лист 'Меню' не содержит данных"))


def _parse_tariff_sheet(ws, data: MenuData):
    headers = _validate_headers(ws, REQUIRED_TARIFF_COLUMNS, "Тарифы", data)
    if any(e.sheet == "Тарифы" for e in data.errors):
        return

    col_idx = {h: i for i, h in enumerate(headers)}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        vals = list(row)
        if all(v is None for v in vals):
            continue

        def get(col):
            idx = col_idx.get(col)
            if idx is None or idx >= len(vals):
                return None
            return vals[idx]

        program = str(get("Программа") or "").strip()
        if program not in VALID_PROGRAMS:
            data.errors.append(ParseError("Тарифы", row_num, "Программа", f"Недопустимая программа '{program}'"))
            continue

        try:
            people = int(get("Люди"))
        except (ValueError, TypeError):
            data.errors.append(ParseError("Тарифы", row_num, "Люди", "Значение должно быть целым числом"))
            continue

        try:
            price = float(get("Цена_за_неделю"))
            if price < 0:
                raise ValueError()
        except (ValueError, TypeError):
            data.errors.append(ParseError("Тарифы", row_num, "Цена_за_неделю", "Значение должно быть >= 0"))
            continue

        data.tariffs.append(Tariff(Программа=program, Люди=people, Цена_за_неделю=price))


def _parse_settings_sheet(ws, data: MenuData):
    headers = _validate_headers(ws, REQUIRED_SETTING_COLUMNS, "Настройки", data)
    if any(e.sheet == "Настройки" for e in data.errors):
        return

    col_idx = {h: i for i, h in enumerate(headers)}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        vals = list(row)
        if all(v is None for v in vals):
            continue

        def get(col):
            idx = col_idx.get(col)
            if idx is None or idx >= len(vals):
                return None
            return vals[idx]

        key = str(get("Ключ") or "").strip()
        val = str(get("Значение") or "").strip()
        if key:
            data.settings[key] = val
