import os
import sys
import pytest

PROJECT_ROOT = r'C:\Users\Ольга\OneDrive\Dokumente\Food delivery service'
sys.path.insert(0, PROJECT_ROOT)

SAMPLE_PATH = os.path.join(PROJECT_ROOT, "sample_data", "sample_menu_ru_headers.xlsx")

from backend.models.schemas import PersonGroup


def test_parse_sample_menu():
    from backend.menu_parser import parse_menu
    data = parse_menu(SAMPLE_PATH)
    assert len(data.errors) == 0, f"Errors: {[str(e) for e in data.errors]}"
    assert len(data.menu_items) > 0
    assert len(data.tariffs) > 0
    assert len(data.settings) > 0


def test_menu_items_have_valid_programs():
    from backend.menu_parser import parse_menu, VALID_PROGRAMS
    data = parse_menu(SAMPLE_PATH)
    for item in data.menu_items:
        assert item.Программа in VALID_PROGRAMS


def test_menu_items_have_valid_meals():
    from backend.menu_parser import parse_menu, VALID_MEALS
    data = parse_menu(SAMPLE_PATH)
    for item in data.menu_items:
        assert item.Приём in VALID_MEALS


def test_menu_days_in_range():
    from backend.menu_parser import parse_menu
    data = parse_menu(SAMPLE_PATH)
    for item in data.menu_items:
        assert 1 <= item.День <= 7


def test_calc_basic():
    from backend.menu_parser import parse_menu
    from backend.calculator import calculate
    data = parse_menu(SAMPLE_PATH)
    groups = [PersonGroup(program="Classic", adults=2, children=0)]
    result = calculate(data, groups, 7, 1)
    assert len(result.groups) == 1
    g = result.groups[0]
    assert g.program == "Classic"
    assert g.adults == 2
    assert g.children == 0
    assert g.people_equiv == 2.0
    assert result.days == 7
    assert result.total_kit_price > 0
    assert result.per_person_per_day > 0
    assert len(g.plan) == 7


def test_calc_with_children():
    from backend.menu_parser import parse_menu
    from backend.calculator import calculate
    data = parse_menu(SAMPLE_PATH)
    groups = [PersonGroup(program="Classic", adults=1, children=2)]
    result = calculate(data, groups, 5, 1)
    g = result.groups[0]
    expected_people = 1 + 2 * 0.5
    assert g.people_equiv == expected_people
    assert len(g.plan) == 5


def test_kit_override():
    from backend.menu_parser import parse_menu
    from backend.calculator import calculate
    data = parse_menu(SAMPLE_PATH)

    groups2 = [PersonGroup(program="Classic", adults=2, children=0)]
    result = calculate(data, groups2, 7, 1)
    assert result.total_kit_price == 6930

    groups3 = [PersonGroup(program="Classic", adults=3, children=0)]
    result = calculate(data, groups3, 7, 1)
    assert result.total_kit_price == 10395


def test_kit_override_no_match():
    from backend.menu_parser import parse_menu
    from backend.calculator import calculate
    data = parse_menu(SAMPLE_PATH)
    groups = [PersonGroup(program="Classic", adults=5, children=0)]
    result = calculate(data, groups, 7, 1)
    g = result.groups[0]
    assert result.total_kit_price == g.computed_price


def test_per_person_per_day():
    from backend.menu_parser import parse_menu
    from backend.calculator import calculate
    data = parse_menu(SAMPLE_PATH)
    groups = [PersonGroup(program="Balance", adults=2, children=0)]
    result = calculate(data, groups, 7, 1)
    expected_ppd = round(7812 / 2 / 7, 1)
    assert result.per_person_per_day == expected_ppd


def test_partial_days():
    from backend.menu_parser import parse_menu
    from backend.calculator import calculate
    data = parse_menu(SAMPLE_PATH)
    groups = [PersonGroup(program="Vegan", adults=1, children=0)]
    result = calculate(data, groups, 3, 2)
    g = result.groups[0]
    assert len(g.plan) == 3
    days_in_plan = [p.День for p in g.plan]
    assert days_in_plan == [2, 3, 4]


def test_day_wraps_around():
    from backend.menu_parser import parse_menu
    from backend.calculator import calculate
    data = parse_menu(SAMPLE_PATH)
    groups = [PersonGroup(program="Classic", adults=1, children=0)]
    result = calculate(data, groups, 3, 6)
    g = result.groups[0]
    days_in_plan = [p.День for p in g.plan]
    assert days_in_plan == [6, 7, 1]


def test_discount_7_days():
    from backend.menu_parser import parse_menu
    from backend.calculator import calculate
    data = parse_menu(SAMPLE_PATH)
    groups = [PersonGroup(program="Classic", adults=2, children=0)]
    result = calculate(data, groups, 7, 1)
    assert result.discount_percent == 10.0
    assert result.discount_amount == 693.0
    assert result.final_price == 6237.0
    assert result.per_day_cost == round(6930 / 7, 1)


def test_no_discount_less_than_7_days():
    from backend.menu_parser import parse_menu
    from backend.calculator import calculate
    data = parse_menu(SAMPLE_PATH)
    groups = [PersonGroup(program="Classic", adults=2, children=0)]
    result = calculate(data, groups, 5, 1)
    assert result.discount_percent == 0.0
    assert result.discount_amount == 0.0
    assert result.final_price == result.total_kit_price


def test_per_day_cost():
    from backend.menu_parser import parse_menu
    from backend.calculator import calculate
    data = parse_menu(SAMPLE_PATH)
    groups = [PersonGroup(program="Balance", adults=2, children=0)]
    result = calculate(data, groups, 7, 1)
    expected_per_day = round(7812 / 7, 1)
    assert result.per_day_cost == expected_per_day


def test_multi_group():
    from backend.menu_parser import parse_menu
    from backend.calculator import calculate
    data = parse_menu(SAMPLE_PATH)
    groups = [
        PersonGroup(program="Classic", adults=2, children=0),
        PersonGroup(program="Balance", adults=1, children=1),
    ]
    result = calculate(data, groups, 7, 1)
    assert len(result.groups) == 2
    assert result.total_people == 4
    assert result.total_kit_price > 0
    assert result.discount_percent == 10.0


def test_parse_invalid_program():
    from backend.menu_parser import parse_menu
    import openpyxl
    import tempfile

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Меню"
    ws.append(["Программа", "День", "Приём", "Блюдо_ID", "Название_блюда",
               "Описание", "Порции_на_человека", "Цена_за_порцию", "Вес_грамм", "Теги"])
    ws.append(["Invalid", 1, "Завтрак", "X1", "Test", "Desc", 1, 100, 200, ""])

    ws2 = wb.create_sheet("Тарифы")
    ws2.append(["Программа", "Люди", "Цена_за_неделю"])

    ws3 = wb.create_sheet("Настройки")
    ws3.append(["Ключ", "Значение"])

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb.save(f.name)
        path = f.name

    try:
        data = parse_menu(path)
        assert len(data.errors) > 0
    finally:
        os.unlink(path)
